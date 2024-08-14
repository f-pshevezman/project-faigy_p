from reportlab.pdfgen import canvas
from reportlab.lib import pagesizes
import json
from flask import Flask, request,jsonify
import openpyxl
import matplotlib.pyplot as plt
import os
import sys

script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
app = Flask(__name__)
@app.route('/upload_excel_file', methods=['POST'])
def upload_fie():
    if 'file' not in request.files:
        return "No file part in the request"
    file = request.files['file']
    if file.filename == '':
        return "No selected file"
    file.save('C:/faigypshevezman/project-python/Uploads/'+file.filename)

    received_workbook = openpyxl.load_workbook('C:/faigypshevezman/project-python/Uploads/'+file.filename)
    num_sheets = len(received_workbook.sheetnames)

    response={
        'file_path':'C:/faigypshevezman/project-python/Uploads/'+file.filename ,
        'num_sheets': num_sheets
    }
    print("The file has been uploaded")
    return response
@app.route('/build_report',methods=['POST'])
def build_report():
    data = request.get_json()
    return build_report(data)
def build_report(data):
    file_path=data.get('file_path')
    sheets=data.get('sheets')
    report_data = []
    wb = openpyxl.load_workbook(file_path)

    for sheet_info in sheets:
        sheet_name = sheet_info.get('sheet')
        operation = sheet_info.get('operation')
        columns = sheet_info.get('columns')
        ws = wb[sheet_name]
        if operation == 'sum':
            for column in columns:
                result=0
                column_data = [cell.value for cell in ws[column]]
                result = sum([value for value in column_data if isinstance(value, (int, float))])
                report_data.append({"sheetname":sheet_name,"column":column,"sum":result})
        elif operation == 'average':
            for column in columns:
                result=0
                column_data = [cell.value for cell in ws[column]]
                result = sum([value for value in column_data if isinstance(value, (int, float))])/len(column_data)
                report_data.append({"sheetname": sheet_name, "column": column, "average": result})
    pdf_function(report_data)
    jsonify(report_data)
    return ''
def pdf_function(report):
    pdf_file = "output.pdf"
    width, height = pagesizes.A4

    c = canvas.Canvas(pdf_file, pagesize=pagesizes.A4)
    c.setFont("Helvetica", 12)

    formatted_data = json.dumps(report, indent=4)

    c.drawString(50, height - 50, "JSON Data:")
    c.drawString(50, height - 70, formatted_data)
    c.save()
    print(f'PDF report generated and saved to {pdf_file}')

def count_sheets_in_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_count = len(workbook.sheetnames)
    print("Number of sheets in the Excel file:", sheet_count)
    return sheet_count

# file_name="excel-files/file1.xlsx"
# full_path=os.path.join(project_dir,file_name)
# print("num of sheets:",count_sheets_in_excel_file(full_path))

def calculate_sum_of_fields(path_file):
    total_sum = 0
    workbook = openpyxl.load_workbook(filename=path_file)

    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]

        for row in current_sheet.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    total_sum += cell_value
    print("The total sum is", total_sum)
    return total_sum

# Example usage
file_name="excel-files/file1.xlsx"
full_path=os.path.join(script_dir,file_name)
print("Sum of fields:",calculate_sum_of_fields(full_path))

def presentation_column_graph(path_file):
    workbook = openpyxl.load_workbook(filename=path_file)
    x=[]
    y=[]
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        x.append(sheet.title())
        total_sum = 0
        for row in current_sheet.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    total_sum += cell_value
        y.append(total_sum)
    plt.bar(x, y)
    plt.xlabel('X-Excel file sheets')
    plt.ylabel('Y-Sum of each sheet')
    plt.title('Column Graph')
    plt.savefig('sum_fields')
    plt.show()

file_name="excel-files/file1.xlsx"
full_path=os.path.join(script_dir,file_name)
presentation_column_graph(full_path)

def average_excel_files(file_path_list):
    i=1
    x=[]
    y=[]
    for file_path in file_path_list:
        x.append(os.path.basename(file_path))
        y.append(calculate_sum_of_fields(file_path)/count_sheets_in_excel_file(file_path))
        i += 1
    plt.bar(x, y)
    plt.xlabel('X-Excel files')
    plt.ylabel('Y-How many fields')
    plt.title('Average excel files ')
    plt.savefig('average_excel_files.png')
    plt.show()

file_name1="excel-files/file1.xlsx"
full_path1=os.path.join(script_dir,file_name1)
file_name2="excel-files/file2.xlsx"
full_path2=os.path.join(script_dir,file_name2)
average_excel_files([full_path1,full_path2])

def amount_per_sheet(file_path,sheet):
    workbook = openpyxl.load_workbook(file_path)
    current_sheet = workbook[sheet]
    total_sum=0
    for row in current_sheet.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value is not None:
                total_sum += cell_value
    print("The total sum is", total_sum)
    return total_sum
@app.route('/create_pdf_report',methods=['POST'])
def create_pdf_report():
    data=request.get_json()
    file_path=data.get('file_path')
    if not file_path:
        return "No file path part in the request"
    report_data = []
    i=0
    received_workbook = openpyxl.load_workbook(file_path)
    for sheet in received_workbook.sheetnames:
        report_data.append({'name':'','sheet_name':'','amount_per_sheet':0})
        report_data[i]['name']= os.path.basename(file_path)
        report_data[i]['sheet_name'] = sheet.title()
        report_data[i]['amount_per_sheet']= amount_per_sheet(file_path,sheet)
        i+=1
    report_data.append({'average_excel_files':average_excel_files([file_path])})
    pdf_function(report_data)
    return "the convert succeeded"

if __name__ == '__main__':
    app.run(debug=True)